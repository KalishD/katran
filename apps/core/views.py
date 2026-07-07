from django.shortcuts import render
from django.conf import settings
from django.views.decorators.http import require_GET
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.sites.models import Site
from django.urls import reverse
from django.contrib.sitemaps import Sitemap
from django.contrib.sitemaps import views as sitemap_views
from django.urls import get_resolver
import csv, codecs, os, re, operator, io, requests
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU # Optional, for precise sizing
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.styles import Alignment # For cell content alignment, if needed
from openpyxl.styles import Alignment  


from apps.store.models import Product, Category, Brand
from apps.blog.models import Post
from django.views.generic import TemplateView
from katran.sitemaps import StaticViewSitemap, CategorySitemap, ProductSitemap, BrandSitemap, PostsSitemap

import json
import logging
from django.views.decorators.csrf import csrf_exempt


# ── Лендинги: Импортозамещение, Разработка под заказ ──

def import_substitution(request):
    """Страница «Подберём аналог» — импортозамещение."""
    description = 'Подберём аналог пневматического инструмента Atlas Copco, Chicago Pneumatic, Desoutter, Ingersoll Rand, Fuji, Sumake. Российские замены для импортного пневмоинструмента.'
    keywords = 'импортозамещение пневмоинструмент, аналог Atlas Copco, замена Chicago Pneumatic, российский пневмоинструмент'
    context = {
        'description': description,
        'keywords': keywords,
    }
    return render(request, 'import_substitution.html', context)


def custom_development(request):
    """Страница «Разработка под заказ»."""
    description = 'Разработаем и произведём промышленный пневматический инструмент под вашу технологию. Собственная конструкторская база, КБ, испытания, серийное производство.'
    keywords = 'разработка пневмоинструмента под заказ, производство пневматического оборудования, OEM пневмоинструмент, специальный пневмоинструмент'
    context = {
        'description': description,
        'keywords': keywords,
    }
    return render(request, 'custom_development.html', context)


logger = logging.getLogger(__name__)

@csrf_exempt  # CSP-отчёты не содержат CSRF-токен
def csp_report(request):
    if request.method != 'POST':
        return HttpResponse(status=405)  # Method Not Allowed

    try:
        report = json.loads(request.body)
        logger.warning("CSP Violation: %s", report)
        # Дополнительно можно сохранить отчёт в БД или отправить в Slack
    except json.JSONDecodeError:
        return HttpResponse(status=400)  # Bad Request


    return HttpResponse(status=204)  # No Content

# robots.txt
class RobotsTxtView(TemplateView):
    template_name = "robots.txt"

# llms.txt
class LLMSTxtView(TemplateView):
    template_name = "llms.txt"

# YML.yml
class YMLView(TemplateView):
    template_name = "YML.yml"

def html_sitemap(request):
    # Словарь классов Sitemap
    sitemap_classes = {
        'static': StaticViewSitemap,
        'product': ProductSitemap,
        'category': CategorySitemap,
        'brand': BrandSitemap,
        'post': PostsSitemap
    }

    all_urls = []
    sitemap_sections = {} # Для организации URL-адресов по разделам в шаблоне

    current_site = Site(domain='katran-pnevmo.ru', name='katran-pnevmo.ru')
    for name, SitemapClass in sitemap_classes.items():
        # Создаем экземпляр класса Sitemap
        sitemap_instance = SitemapClass()
        
        # Используем метод get_urls() экземпляра для получения всех URL-адресов
        # Этот метод уже вызывает items(), location(), lastmod() и т.д.
        urls_for_section = sitemap_instance.get_urls(site=current_site)
        
        all_urls.extend(urls_for_section)
        sitemap_sections[name.capitalize()] = urls_for_section # Группируем для шаблона

    # Добавляем пагинацию, как было предложено в отчете, для больших карт сайта
    paginator = Paginator(all_urls, 100) # По 100 URL-адресов на странице

    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        # Если номер страницы не целое число, показываем первую страницу
        page_obj = paginator.page(1)
    except EmptyPage:
        # Если страница выходит за пределы диапазона, показываем последнюю страницу
        page_obj = paginator.page(paginator.num_pages)
    description = "Карта сайта katran-pnevmo.ru "
    context = {
        'page_obj': page_obj, # Передаем объект страницы для итерации в шаблоне
        'sitemap_sections': sitemap_sections, # Передаем разделы для структурированного отображения
        'description': description,
    }
    return render(request, 'sitemap.html', context)


from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def temp(request):
    mp011 = Product.objects.filter(sku = 5882).first()
    context = {'mp011': mp011}
    return render(request, 'temp.html', context)


def error_404_view(request):
    main_brand_id = getattr(settings, 'MAIN_BRAND_ID', 1)
    katran_products = Product.objects.filter(brand=main_brand_id, category__main_category__in=range(1, 2), is_visible=True).order_by('sku')
    description = 'Эта страница не найдена, но у нас много другого инструмента в каталоге.'
    context = {
        'katran_products': katran_products,
        'description': description,
    }
    return render(request, '404.html', context)

def error_500_view(request):
    main_brand_id = getattr(settings, 'MAIN_BRAND_ID', 1)
    katran_products = Product.objects.filter(brand=main_brand_id, category__main_category__in=range(1, 2), is_visible=True).order_by('sku')
    description = 'Ой, что-то с нашим сервером. Но мы скоро всё починим.'

    context = {
        'katran_products': katran_products,
        'description': description,
    }
    return render(request, '500.html', context)



def frontpage(request):
    main_brand_id = getattr(settings, 'MAIN_BRAND_ID', 1)
    katran_products = Product.objects.filter(brand=main_brand_id, category__main_category__in=range(1, 2), is_visible=True).order_by('category')
    cangairgrinders = Product.objects.filter(category__in=getattr(settings, 'AIR_GRINDER_CATEGORIES', [16]), is_visible=True).order_by('sku')
    anglegrinders = Product.objects.filter(category__in=getattr(settings, 'ANGLE_GRINDER_CATEGORIES', [15]), is_visible=True).order_by('-price')
    airhammers = Product.objects.filter(category__in=range(5, 7), is_visible=True).order_by('-price')
    posts = Post.objects.all()
    keywords = 'Купить пневмоинструмент, пневматические молотки отбойные, пневматические молотки рубильные, российский пневмоинструмент'
    description = 'Магазин пневматического инструмента, ООО "Катран-Пневмо" более 30 лет на рынке. Купить пневмоинструмент в СПб.'
    context = {
        'katran_products': katran_products,
        'cangairgrinders': cangairgrinders,
        'anglegrinders': anglegrinders,
        'airhammers': airhammers,
        'posts': posts,
        'keywords': keywords,
        'description': description,
    }

    return render(request, 'frontpage.html', context)

def politics(request):
    keywords = ''
    description = 'Политика обработки персональных данных'
    context = {
        'keywords': keywords,
        'description': description,
    }
    return render(request, 'politics.html', context)

def politics_agree(request):
    keywords = ''
    description = 'Согласие с политикой персональных данных'
    context = {
        'keywords': keywords,
        'description': description,
    }
    return render(request, 'politics_agree.html', context)

def production(request):
    mp006 = Product.objects.filter(sku = 4).first()
    mp01122 = Product.objects.filter(sku = 5882).first()
    rm8 = Product.objects.filter(sku = 17).first()
    rm12 = Product.objects.filter(sku = 18).first()
    rm16 = Product.objects.filter(sku = 19).first()
    tp28a = Product.objects.filter(sku = 31).first()
    tpv3a = Product.objects.filter(sku = 237).first()
    mps = Product.objects.filter(sku = 1).first()
    ppf420 = Product.objects.filter(sku = 15).first()
    pvm = Product.objects.filter(sku = 774).first()
    ru64 = Product.objects.filter(sku = 3397).first()
    mp011_list = [mp01122,]
    rm_list = [rm8,rm12,rm16]
    tramb_list = [tp28a, tpv3a]
    pvm_list = [ppf420, pvm, ru64]
    keywords = ''
    description = 'Пневматический инструмент российского производства. Шлифовальные машины, пневмомолотки и трамбовки от производителя в Санкт-Петербурге. '
    context = {
        'mp006': mp006,
        'mp01122': mp01122,
        'rm8': rm8,
        'rm12': rm12,
        'rm16': rm16,
        'mps': mps,
        'mp011_list': mp011_list,
        'rm_list': rm_list,
        'tramb_list': tramb_list,
        'pvm_list': pvm_list,
        'keywords': keywords,
        'description': description,
    }

    return render(request, 'production.html', context)


def about(request):
    keywords = 'О компании Катран-Пневмо, производитель пневматического инструмента, патенты, история компании, пневмоинструмент Санкт-Петербург, пневматические шлифмашин гайковерты молотки'
    description = 'ООО «Катран-Пневмо» — производитель и поставщик промышленного пневматического инструмента с 1990 года. Собственное производство в СПб, стендовые испытания, авторизованный дистрибьютор U-Tools и Daewoo. Гарантия 12 месяцев.'
    context = {
        'keywords': keywords,
        'description': description,
    }
    return render(request, 'about.html', context)

def contacts(request):
    keywords = 'Контакты Катран-Пневмо, телефон, адрес, реквизиты, Санкт-Петербург'
    description = 'Контакты ООО "Катран-Пневмо": телефон +7 (812) 331 79 09, адрес в Санкт-Петербурге, банковские реквизиты, график работы.'
    context = {
        'keywords': keywords,
        'description': description,
    }
    return render(request, 'contacts.html', context)


def sale_price(request):
    products = Product.objects.filter(is_in_sales_price = True).order_by('title')
    categories = Category.objects.filter(products__is_in_sales_price=True).distinct().order_by('ordering', '-id')
    keywords = ''
    description = 'Прайс-лист пневматического инструмента фирмы Jesda'
    context = {
        'categories': categories,
        'products': products,
        'keywords': keywords,
        'description': description,
    }
    return render(request, 'sale_price.html', context)


def export_pricelist(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 100 
    ws.column_dimensions['C'].width = 20 
    colA = ws.column_dimensions['A']
    colB = ws.column_dimensions['B']
    colC = ws.column_dimensions['C']

    IMG_WIDTH_PX = 80
    IMG_HEIGHT_PX = 80
    # Ширина колонки 'A' установлена в 20.

    products = Product.objects.filter(is_in_sales_price = True).order_by('title')
    categories = Category.objects.filter(products__is_in_sales_price=True).distinct().order_by('ordering', '-id')
    counter = 0
    for category in categories:
      counter += 1
      ws.append([category.title,'',''])
      print(counter)
      print('Fill cell')
      ws.merge_cells(start_row=counter, start_column=1, end_row=counter, end_column=3)  
      cell_index = 'A'+str(counter)
      cell = ws[cell_index]  
      cell.fill = PatternFill(start_color='dae9f6', end_color='dae9f6', fill_type='solid')
      cell.font = Font(bold=True, size=16)
      for product in  Product.objects.filter(is_in_sales_price = True, category = category.id).order_by('title'):
        counter += 1
        ws.row_dimensions[counter].height = 75
        
        product_var = '\n'
        # f"My name is {name} and I'm {age} years old"
        for varitem in product.variable_set.all():
          product_var += f"{varitem}: {varitem.value}{varitem.varitem.dimention if varitem.varitem.dimention else ''}; "
        ws.append(['',product.title + '\n' + product_var, str(product.price) + ' с НДС'])
        if product.image:
          try:
            image_path = product.image.path 
            image = Image(image_path)
            
            # Установка размеров изображения
            image.width = IMG_WIDTH_PX
            image.height = IMG_HEIGHT_PX
            
            # Use TwoCellAnchor with offsets for precise positioning and sizing (do some experiment)
            offset_x = offset_y = 30000  # Example offset values for positioning
            _from = AnchorMarker(col=0, row=counter - 1, colOff=offset_x, rowOff=offset_y)
            to = AnchorMarker(col=1, row=counter, colOff=-offset_x, rowOff=-offset_y)
            image.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=to)

            ws.add_image(image)
              
          except FileNotFoundError:
            print(f"FileNotFoundError: Файл изображения не найден по пути {product.image.path}")
            pass
          except Exception as e:
            print(f"Error adding image for product {product.sku}: {e}")
            pass
          ws['B'+str(counter)].alignment = Alignment(wrap_text=True) 
      ws.append(['','',''])
      counter += 1

    colA.font = Font(name='Calibri', size=14, color="09131f")
    colB.font = Font(name='Calibri', size=14, color="09131f")
    colC.font = Font(name='Calibri', size=16, color="09131f")

    buffer = io.BytesIO()  
    wb.save(buffer)  
    buffer.seek(0)  

    # Формируем HttpResponse  
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")  
    response["Content-Disposition"] = "attachment; filename=JSD-Tools_katran-pnevmo.xlsx"  

    return response 