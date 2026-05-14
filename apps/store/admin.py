from django.contrib import admin
from django.core.files import File
from django.urls import reverse
from django.utils.http import urlencode
from django.utils.html import format_html
from django import forms
from django.urls import path
from django.shortcuts import render, redirect
from slugify import slugify
import csv, codecs, os, re, operator, io, requests
from io import BytesIO
from PIL import Image
from apps.store.models import Category, Product, Brand, Variable, VariableItem, MainCategory
from urllib.parse import parse_qsl, urljoin, urlparse
from urllib.request import urlopen
from apps.core.utils import *
from import_export.admin import ExportActionMixin
from django.http import HttpResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU # Optional, for precise sizing
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.styles import Alignment # For cell content alignment, if needed
from openpyxl.styles import Alignment  

from django import forms
from django.contrib.admin.widgets import FilteredSelectMultiple

from django_summernote.admin import SummernoteModelAdmin
from django.contrib.sites.models import Site

class CustomSiteAdmin(admin.ModelAdmin):
    list_display = ('domain', 'name')

class VariableInline(admin.TabularInline):
  model = Variable
  raw_id_fields = ['product']

# CSV Import Products
class CsvImportForm(forms.Form):
    csv_files = forms.FileField()

# CSV Import Products
class CsvImportProductForm(forms.Form):
    csv_files = forms.FileField()
    csv_files_1 = forms.FileField()



@admin.register(VariableItem)
class VariableItemAdmin(admin.ModelAdmin):
  change_list_template = "variableitem_changelist.html"
  list_display = ("title","dimention", "is_primary")
  search_fields = ("title__contains",)
  fields = ("title","dimention", "is_primary")

  def get_urls(self):
    urls = super().get_urls()
    my_urls = [
      path('import-csv-var/', self.import_csv),
    ]
    return my_urls + urls
    
  def import_csv(self, request):
    if request.method == "POST":
      csv_file = request.FILES["csv_files"]
      with io.TextIOWrapper(csv_file, encoding="utf-8") as text_file:
        reader = csv.reader(text_file, lineterminator='\n', delimiter=';')
        for row in reader:
          try:
            varitem = VariableItem.objects.get_or_create(
              id = row[0],
              title = row[1],
              dimention = row[2],
            )
          except Exception as inst:
            print(inst)
            print('Exception:',row)
        self.message_user(request, "Your csv file has been imported")
        return redirect("..")
    form = CsvImportForm()
    payload = {"form": form}
    return render(
        request, "csv_form.html", payload
    )

class ProductAdminForm(forms.ModelForm):
    class Meta:
        model = Product
        fields = '__all__'

    parts = forms.ModelMultipleChoiceField(
        queryset=Product.objects.none(),  # заполняем в __init__
        required=False,
        widget=FilteredSelectMultiple('Запчасти', False)
    )
    analogs = forms.ModelMultipleChoiceField(
        queryset=Product.objects.none(),
        required=False,
        widget=FilteredSelectMultiple('Аналоги', False)
    )
    similar_products = forms.ModelMultipleChoiceField(
        queryset=Product.objects.none(),
        required=False,
        widget=FilteredSelectMultiple('Похожие товары', False)
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # базовый queryset — все продукты
        qs = Product.objects.filter(is_visible=True)
        
        # parts: только main_category_id в [3,4,5]
        self.fields['parts'].queryset = qs.filter(category__main_category_id__in=[3,4,5])
        
        # analogs: только main_category_id = 1
        self.fields['analogs'].queryset = qs.filter(category__main_category_id=1)
        
        # similar: тот же main_category, что и у текущего объекта
        if self.instance and self.instance.pk:
            mc = self.instance.category.main_category
            self.fields['similar_products'].queryset = qs.filter(category__main_category=mc)
        else:
            # при создании объекта — пустой список или весь qs, как вам удобнее
            self.fields['similar_products'].queryset = Product.objects.none()

# admin.site.register(Product)
@admin.register(Product)
class ProductAdmin(ExportActionMixin, SummernoteModelAdmin, admin.ModelAdmin):
  
  change_list_template = "products_changelist.html"

  def get_urls(self):
    urls = super().get_urls()
    my_urls = [
        path('import-csv/', self.import_csv),
        path('import-csv-price/', self.import_csv_price),
        path('export-csv-price/', self.export_pricelist),
        path('set-price-without-taxes/', self.set_wo_tax_price),
    ]
    return my_urls + urls
  
  def set_wo_tax_price(self, request):
    products = Product.objects.all()  
    for product in products:
      if product.price != 0:
        product.in_stock = 1
      else:
        product.in_stock = 0
      # product.price = product.price_wo_tax * (1 + product.tax/100)
       # product.price_wo_tax = product.price / (1 + product.tax/100)
      product.save()
    return redirect("..")
  
  def export_csv_price(self, request):
    # Создаём рабочую книгу и лист  
    wb = Workbook()  
    ws = wb.active  
    ws.title = "Products price"  
    # Пишем заголовки в первой строке  
    ws.append(["sku",
               "title",
               "price",
               "is_visible"])  
    # Извлекаем нужные данные и добавляем их построчно  
    products = Product.objects.all()  
    for product in products:  
        ws.append([product.sku, product.title, product.price, product.is_visible])  
    # Сохраняем в память  
    buffer = io.BytesIO()  
    wb.save(buffer)  
    buffer.seek(0)  
    # Формируем HttpResponse  
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")  
    response["Content-Disposition"] = "attachment; filename=products.xlsx"  
    return response 

  def export_pricelist(self, request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 100 
    ws.column_dimensions['C'].width = 20 
    colA = ws.column_dimensions['A']
    colB = ws.column_dimensions['B']
    colC = ws.column_dimensions['C']

    IMG_WIDTH_PX = 80
    IMG_HEIGHT_PX = 80
    # Ширина колонки 'A' установлена в 20.

    products = Product.objects.filter(is_in_sales_price = True).order_by('title')
    categories = Category.objects.filter(products__is_in_sales_price=True).distinct().order_by('ordering', '-id')
    counter = 0
    for category in categories:
      counter += 1
      ws.append([category.title,'',''])
      print(counter)
      print('Fill cell')
      ws.merge_cells(start_row=counter, start_column=1, end_row=counter, end_column=3)  
      cell_index = 'A'+str(counter)
      cell = ws[cell_index]  
      cell.fill = PatternFill(start_color='dae9f6', end_color='dae9f6', fill_type='solid')
      cell.font = Font(bold=True, size=16)
      for product in  Product.objects.filter(is_in_sales_price = True, category = category.id).order_by('title'):
        counter += 1
        ws.row_dimensions[counter].height = 75
        
        product_var = '\n'
        # f"My name is {name} and I'm {age} years old"
        for varitem in product.variable_set.all():
          product_var += f"{varitem}: {varitem.value}{varitem.varitem.dimention if varitem.varitem.dimention else ''}; "
        ws.append(['',product.title + '\n' + product_var, str(product.price) + ' с НДС'])
        if product.image:
          try:
            image_path = product.image.path 
            image = Image(image_path)
            
            # Установка размеров изображения
            image.width = IMG_WIDTH_PX
            image.height = IMG_HEIGHT_PX
            
            # Use TwoCellAnchor with offsets for precise positioning and sizing (do some experiment)
            offset_x = offset_y = 30000  # Example offset values for positioning
            _from = AnchorMarker(col=0, row=counter - 1, colOff=offset_x, rowOff=offset_y)
            to = AnchorMarker(col=1, row=counter, colOff=-offset_x, rowOff=-offset_y)
            image.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=to)

            ws.add_image(image)
              
          except FileNotFoundError:
            print(f"FileNotFoundError: Файл изображения не найден по пути {product.image.path}")
            pass
          except Exception as e:
            print(f"Error adding image for product {product.sku}: {e}")
            pass
          ws['B'+str(counter)].alignment = Alignment(wrap_text=True) 
      ws.append(['','',''])
      counter += 1

    colA.font = Font(name='Calibri', size=14, color="09131f")
    colB.font = Font(name='Calibri', size=14, color="09131f")
    colC.font = Font(name='Calibri', size=16, color="09131f")

    buffer = io.BytesIO()  
    wb.save(buffer)  
    buffer.seek(0)  

    # Формируем HttpResponse  
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")  
    response["Content-Disposition"] = "attachment; filename=JSD-Tools_katran-pnevmo.xlsx"  

    return response 
    
  def import_csv_price(self, request):
    if request.method == "POST":
      csv_file = request.FILES["csv_files"]
      with io.TextIOWrapper(csv_file, encoding='cp1251') as text_file:
        reader = csv.reader(text_file, lineterminator='\n', delimiter=';')
        for count, row in enumerate(reader, start=1):
          try:
            product = Product.objects.filter(id=int(row[0])).first()
            product.price = float(row[2].replace(',', '.'))
            # product.title = str(row[1])
            # product.slug = slugify(str(row[1]))
            product.is_visible = int(row[3])
            product.save()
            # product = Product.objects.update_or_create(
            #   id = int(row[0]),
            #   sku = int(row[0]),
            #   title = str(row[1]),
            #   slug = slugify(str(row[1])),
            #   description = '',
            #   price = float(row[2].replace(',', '.')),
            #   is_visible = int(row[3]),
            #   )
          except Exception as inst:
              print('Error, line #', count)
              print(inst)
              print(row[0], row[1])
        self.message_user(request, "Your csv file has been imported")
        return redirect("..")
    form = CsvImportForm()
    payload = {"form": form}
    return render(
      request, "csv_form_price.html", payload
    )
  
  
  def import_csv(self, request):
    if request.method == "POST":
        csv_file = request.FILES["csv_files"]
        with io.TextIOWrapper(csv_file, encoding="utf-8") as text_file:
          reader = csv.reader(text_file, lineterminator='\n', delimiter = ';')
          for count, row in enumerate(reader, start=1):
            # print('Product #',count, type(count))
            # print('!!!_row_!!!',row)
            i = 7
            try:
              slug = slugify(str(row[1]))
              if row[0] == '2865':
                print('DAG')
              product = Product.objects.update_or_create(
                id = int(row[0]),
                sku = int(row[0]),
                title = row[1],
                slug = slug,
                description = row[5],
                price = float(row[3].replace(',', '.')),
                is_features = 0,
                category_id = int(row[2]) if row[2] != '' else None,
                brand_id = int(row[6]) if row[6] != '' else None,
              )
              if row[4]:
                url = row[4]
                parse_object = urlparse(url)
                domain = parse_object.netloc
                path = parse_object.path
                filename = url.split('/')[-1]
                # print('Domain, path, filename',domain, path, filename)
                response = requests.get(url)
                response_content = BytesIO(urlopen(url).read())
                pil_image = Image.open(response_content).convert('RGB')
                django_file = pil_to_django(pil_image)
                product[0].image.save(filename, django_file)
              tmp_row = []
              if row[i] != '':
                while row[i] != '': 
                  # tmp_row.extend((row[i],row[i+1]))
                  product[0].variable_set.get_or_create(
                    varitem = VariableItem.objects.filter(id = int(row[i]))[0],
                    value = row[i+1]
                  )
                  i += 2  
            except Exception as inst:
              print('Error, line #', count)
              print(inst)
              print(row[0])
              print(i)
          self.message_user(request, "Your csv file has been imported")
          return redirect("..")
    form = CsvImportForm()
    payload = {"form": form}
    return render(
        request, "csv_form.html", payload
    )
  

    # уже есть:
  list_display = ("sku", "title", "product_category", "brand", "price_wo_tax", "price", "ordering", "is_in_sales_price", "is_features", "is_visible")
  fields = (
      "category", "brand", "sku", "title", "slug", "article", "price_wo_tax", "price", "ordering", 
      "is_features", "is_sale", "is_bestseller", "is_visible", "is_in_sales_price", "keywords", "in_stock", "description", "image",
      "analog", "parts", "similar_products", "partlist"
  )
  
  search_fields = ("title__icontains",)
  prepopulated_fields = {'slug': ('title',) }
  filter_horizontal = ('parts', 'similar_products',)  # удобный выбор через виджет

  def formfield_for_manytomany(self, db_field, request, **kwargs):
      # общий queryset только видимые продукты
      qs = Product.objects.filter(is_visible=True)

      if db_field.name == 'parts':
          # запчасти: только main_category_id в [3,4,5]
          kwargs['queryset'] = qs.filter(category__main_category_id__in=[3,4,5])
      elif db_field.name == 'analogs':
          # аналоги: только main_category_id = 1
          kwargs['queryset'] = qs.filter(category__main_category_id=1)
      elif db_field.name == 'similar_products':
          # похожие: тот же main_category, что у редактируемого объекта
          # request.resolver_match doesn't give obj, поэтому смотрим request.path
          obj_id = request.resolver_match.kwargs.get('object_id')
          if obj_id:
              prod = Product.objects.filter(pk=obj_id).first()
              if prod:
                  mc = prod.category.main_category
                  kwargs['queryset'] = qs.filter(category__main_category=mc)
              else:
                  kwargs['queryset'] = qs.none()
          else:
              kwargs['queryset'] = qs.none()

      return super().formfield_for_manytomany(db_field, request, **kwargs)


  def save_model(self, request, obj, form, change):
      # сначала сохраняем сам объект
      super().save_model(request, obj, form, change)

      # Если у текущего товара выбран аналог
      if obj.analog:
          other = obj.analog
          # и у другого товара аналог не установлен или установлен не на нас
          if other.analog_id != obj.id:
              other.analog = obj
              other.save()
      
      # Если убрали аналог — очищаем у бывшего
      if change:
          old = form.initial.get('analog')
          if old and (not obj.analog or int(old) != obj.analog_id):
              prev = Product.objects.filter(pk=old).first()
              if prev and prev.analog_id == obj.id:
                  prev.analog = None
                  prev.save()


  inlines = [VariableInline]
  save_as = True
  # summernote_fields = ('description',)
  def product_category(self,obj):
    url = (
      reverse("admin:store_product_changelist")
      + "?"
      + urlencode({"category_id": f"{obj.category.id}"})
    )
    return format_html('<a href={}>{}</a>', url, obj.category)



@admin.register(Brand)
class BrandAdmin(admin.ModelAdmin):
  list_display = ("title",'is_on',"product_count","ordering","country")
  fields = ("title","slug","description",'is_on',"ordering","country", "image")
  prepopulated_fields = {'slug': ('title',) }

  def product_count(self, obj):
    count = Product.objects.filter(brand=obj).count()

    url = (
      reverse("admin:store_product_changelist")
      + "?"
      + urlencode({"brand_id": f"{obj.id}"})
    )
    return format_html('<a href={}>{} Products</a>', url, count)

  product_count.short_description = "Products"

# admin.site.register(MainCategory)
@admin.register(MainCategory)
class MainCategoryAdmin(admin.ModelAdmin):
  list_display = ("title",)
  prepopulated_fields = {'slug': ('title',) }
  fields = ("title","slug","ordering")
  
# admin.site.register(Category)
@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
  
  change_list_template = "category_changelist.html"

  def get_urls(self):
    urls = super().get_urls()
    my_urls = [
        path('import-csv/', self.import_csv),
    ]
    return my_urls + urls
  
  def product_count(self, obj):
    count = Product.objects.filter(category=obj).count()

    url = (
      reverse("admin:store_product_changelist")
      + "?"
      + urlencode({"category_id": f"{obj.id}"})
    )
    return format_html('<a href={}>{} Products</a>', url, count)
  
  def get_form(self, request, obj=None, **kwargs):
    form = super().get_form(request, obj, **kwargs)
    form.base_fields["title"].label = "Название: "
    return form

  def import_csv(self, request):
    if request.method == "POST":
        csv_file = request.FILES["csv_files"]
        with io.TextIOWrapper(csv_file, encoding="utf-8") as text_file:
          reader = csv.reader(text_file, lineterminator='\n', delimiter = ';')
          for count, row in enumerate(reader, start=1):
            try:
              print('******')
              print('Category #', count)
              print('-=-=-=-=-=-=-=-=-')
              print('row[0] :', row[0])
              print('-=-=-=-=-=-=-=-=-')
              print('row[1] :', row[1])
              print('-=-=-=-=-=-=-=-=-')
              slug = slugify(str(row[1]))
              print('slug :', slug)
              print('-=-=-=-=-=-=-=-=-')
              print('Main Category ID check :')
              cat_id = int(row[0])
              print('cat_id', cat_id)
              print('-')
              if cat_id in range(1,17):
                print('mcat #1')
                mcat_id = 1
              elif cat_id in range(17,20):
                print('mcat #2')
                mcat_id = 2
              elif cat_id in range(20,23):
                print('mcat #3')
                mcat_id = 3
              elif cat_id in range(23,26):
                print('mcat #4')
                mcat_id = 5
              elif cat_id in range(26,34):
                print('mcat #5')
                mcat_id = 4
                
              print('main_category_id :', mcat_id)
              print('-=-=-=-=-=-=-=-=-')
              print('category create :')

                  
              category = Category.objects.update_or_create(
                id = row[0],
                title = row[1],
                main_category_id = mcat_id, 
                slug = slug,
                description = row[1],
                is_features = 0,
              )
            except Exception as inst:
              print('Error, line #', count)
              print(inst)
          self.message_user(request, "Your csv file has been imported")
          return redirect("..")
    form = CsvImportForm()
    payload = {"form": form}
    return render(
        request, "csv_form.html", payload
    )
      
  list_display = ("title", "id", "main_category", "is_features", "description", "ordering","product_count")
  product_count.short_description = "Products"
  fields = ("title", "main_category", "is_features", "description", "summer_description", "ordering", "slug", "image")
  prepopulated_fields = {'slug': ('title',)
  # summernote_fields = ('summer_description',)
  }
  



# admin.site.register(Variable)
@admin.register(Variable)
class VariableAdmin(admin.ModelAdmin):
  # filter_horizontal = ('product',)
  pass